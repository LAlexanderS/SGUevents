from django.contrib.auth.decorators import login_required
from django.contrib.contenttypes.models import ContentType
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from bookmarks.models import Favorite, Registered, Review
from events_available.models import Events_online, Events_offline
from events_cultural.models import Attractions, Events_for_visiting
from users.telegram_utils import send_message_to_user
from events_cultural.views import submit_review
from django.views.decorators.csrf import csrf_exempt
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from bookmarks.forms import SendMessageForm
from users.telegram_utils import send_notification_with_toggle
from bookmarks.models import Registered
from django.utils.timezone import now
import logging
from django.db.models import Avg
from django.conf import settings
from io import BytesIO
from django.utils.text import slugify
try:
    from docxtpl import DocxTemplate
except Exception:
    DocxTemplate = None
import os
from bookmarks.forms import ExportParticipantsForm
from .forms import ExportLogisticsForm
from events_available.models import EventLogistics
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.utils import timezone


logger = logging.getLogger(__name__)


@login_required
def events_add(request, event_slug):
    event = None
    event_type = None
    try:
        event = Events_online.objects.get(slug=event_slug)
        event_type = 'online'
    except Events_online.DoesNotExist:
        try:
            event = Events_offline.objects.get(slug=event_slug)
            event_type = 'offline'
        except Events_offline.DoesNotExist:
            try:
                event = Attractions.objects.get(slug=event_slug)
                event_type = 'attractions'
            except Attractions.DoesNotExist:
                try:
                    event = Events_for_visiting.objects.get(slug=event_slug)
                    event_type = 'for_visiting'
                except Events_for_visiting.DoesNotExist:
                    pass

    if event and request.user.is_authenticated:
        favorite, created = None, False
        if event_type == 'online':
            favorite, created = Favorite.objects.get_or_create(user=request.user, online=event)
        elif event_type == 'offline':
            favorite, created = Favorite.objects.get_or_create(user=request.user, offline=event)
        elif event_type == 'attractions':
            favorite, created = Favorite.objects.get_or_create(user=request.user, attractions=event)
        elif event_type == 'for_visiting':
            favorite, created = Favorite.objects.get_or_create(user=request.user, for_visiting=event)

        if not created:
            favorite.delete()
            added = False
        else:
            added = True

        return JsonResponse({'added': added, 'event_id': favorite.id})

    return JsonResponse({'error': 'Event not found or user not authenticated'}, status=400)

@login_required
def events_remove(request, event_slug):
    if request.method == 'POST':
        favorite = None
        try:
            event = Events_online.objects.get(slug=event_slug)
            favorite = Favorite.objects.get(user=request.user, online=event)
        except Events_online.DoesNotExist:
            try:
                event = Events_offline.objects.get(slug=event_slug)
                favorite = Favorite.objects.get(user=request.user, offline=event)
            except Events_offline.DoesNotExist:
                try:
                    event = Attractions.objects.get(slug=event_slug)
                    favorite = Favorite.objects.get(user=request.user, attractions=event)
                except Attractions.DoesNotExist:
                    try:
                        event = Events_for_visiting.objects.get(slug=event_slug)
                        favorite = Favorite.objects.get(user=request.user, for_visiting=event)
                    except Events_for_visiting.DoesNotExist:
                        return JsonResponse({'error': 'Event not found'}, status=404)

        if favorite:
            favorite.delete()
            return JsonResponse({'removed': True})

    return JsonResponse({'removed': False, 'error': 'Invalid request method'}, status=400)

@login_required
def favorites(request):
    favorites = Favorite.objects.filter(user=request.user).order_by('-created_timestamp')

    events = []
    for fav in favorites:
        if fav.online:
            events.append(fav.online)
        elif fav.offline:
            events.append(fav.offline)
        elif fav.attractions:
            events.append(fav.attractions)
        elif fav.for_visiting:
            events.append(fav.for_visiting)

    reviews = {}
    for event in events:
        content_type = ContentType.objects.get_for_model(event)
        reviews[event.unique_id] = Review.objects.filter(content_type=content_type, object_id=event.id)

    registered_dict = {
        'online': {},
        'offline': {},
        'attractions': {},
        'for_visiting': {}
    }

    # Получаем все зарегистрированные мероприятия и распределяем их по типам
    for event in events:
        if isinstance(event, Events_online):
            reg_event = Registered.objects.filter(user=request.user, online=event).first()
            if reg_event:
                registered_dict['online'][event.id] = reg_event.id
        elif isinstance(event, Events_offline):
            reg_event = Registered.objects.filter(user=request.user, offline=event).first()
            if reg_event:
                registered_dict['offline'][event.id] = reg_event.id
        elif isinstance(event, Attractions):
            reg_event = Registered.objects.filter(user=request.user, attractions=event).first()
            if reg_event:
                registered_dict['attractions'][event.id] = reg_event.id
        elif isinstance(event, Events_for_visiting):
            reg_event = Registered.objects.filter(user=request.user, for_visiting=event).first()
            if reg_event:
                registered_dict['for_visiting'][event.id] = reg_event.id

    registered_flat = {}
    for subdict in registered_dict.values():
        registered_flat.update(subdict)

    liked_slugs = [event.slug for event in events]

    reviews_avg = {}
    for event in events:
        content_type = ContentType.objects.get_for_model(event)
        avg_rating = Review.objects.filter(
            content_type=content_type,
            object_id=event.id,
            rating__isnull=False
        ).aggregate(Avg('rating'))['rating__avg']
        reviews_avg[event.id] = round(avg_rating, 1) if avg_rating else 0

    print(f'Reviews: {reviews_avg}')

    context = {
        'events': events,
        'reviews': reviews,
        'favorites': favorites,
        'liked': liked_slugs,
        'registered': registered_flat,
        'now': now().date(),
        'name_page': 'Избранные',
        'reviews_avg': reviews_avg,
    }
    return render(request, 'bookmarks/favorites.html', context)

def events_attended(request):
    pass
    return render(request, "bookmarks/events_attended.html")

@login_required
def events_registered(request, event_slug):
    event = None
    event_type = None
    favorites = Favorite.objects.filter(user=request.user)

    try:
        event = Events_online.objects.get(slug=event_slug)
        event_type = 'online'
    except Events_online.DoesNotExist:
        try:
            event = Events_offline.objects.get(slug=event_slug)
            event_type = 'offline'
        except Events_offline.DoesNotExist:
            try:
                event = Attractions.objects.get(slug=event_slug)
                event_type = 'attractions'
            except Attractions.DoesNotExist:
                try:
                    event = Events_for_visiting.objects.get(slug=event_slug)
                    event_type = 'for_visiting'
                except Events_for_visiting.DoesNotExist:
                    pass

    if event and request.user.is_authenticated:
        registered, created = None, False
        if event_type == 'online':
            registered, created = Registered.objects.get_or_create(user=request.user, online=event)
        elif event_type == 'offline':
            registered, created = Registered.objects.get_or_create(user=request.user, offline=event)
        elif event_type == 'attractions':
            registered, created = Registered.objects.get_or_create(user=request.user, attractions=event)
        elif event_type == 'for_visiting':
            registered, created = Registered.objects.get_or_create(user=request.user, for_visiting=event)

        # Отправляем уведомление с гиперссылкой
        if created and request.user.telegram_id:
            from users.telegram_utils import get_event_url, create_event_hyperlink, send_message_to_user_with_toggle_button
            
            # Генерируем URL и создаем гиперссылку
            event_url = get_event_url(event)
            event_hyperlink = create_event_hyperlink(event.name, event_url)
            
            message = f"✅ Вы зарегистрировались на мероприятие: {event_hyperlink}."
            
            # Определяем ссылку на чат участников при наличии
            chat_url = None
            if event_type in ('online', 'offline'):
                try:
                    if getattr(event, 'users_chat_id', None) and getattr(event, 'users_chat_link', None):
                        chat_url = event.users_chat_link
                except AttributeError:
                    chat_url = None
            
            send_message_to_user_with_toggle_button(
                request.user.telegram_id, 
                message, 
                registered.id, 
                registered.notifications_enabled,
                chat_url=chat_url
            )

        event.member.add(request.user)

        if created:
            response_data = {'added': True, 'event_id': registered.id, 'event_slug': event_slug}
            if event_type == 'for_visiting':
                event.place_free -= 1
                event.save(update_fields=['place_free'])
                response_data['place_free'] = event.place_free
            return JsonResponse(response_data)
        else:
            return JsonResponse({'added': False, 'error': 'Уже зарегистрированы'}, status=400)

    return JsonResponse({'error': 'Event not found or user not authenticated'}, status=400)

@login_required
def registered_remove(request, event_id):
    if request.method == 'POST':
        event = get_object_or_404(Registered, id=event_id, user=request.user)
        event_slug = (
            event.for_visiting.slug if event.for_visiting else (
                event.online.slug if event.online else (
                    event.offline.slug if event.offline else (
                        event.attractions.slug if event.attractions else None
                    )
                )
            )
        )
        event_name = (
            event.for_visiting.name if event.for_visiting else (
                event.online.name if event.online else (
                    event.offline.name if event.offline else (
                        event.attractions.name if event.attractions else None
                    )
                )
            )
        )

        if event.for_visiting:
            event.for_visiting.member.remove(request.user)
            event.for_visiting.place_free += 1
            event.for_visiting.save(update_fields=['place_free'])
        elif event.online:
            event.online.member.remove(request.user)
        elif event.offline:
            event.offline.member.remove(request.user)

        event.delete()
        telegram_id = request.user.telegram_id
        if telegram_id:
            message = f"\U0000274C Вы успешно отменили регистрацию на мероприятие: {event_name}"
            send_message_to_user(telegram_id, message)
        
        response_data = {'removed': True, 'event_slug': event_slug, 'event_name': event_name}
        if event.for_visiting:
            response_data['place_free'] = event.for_visiting.place_free
        return JsonResponse(response_data)

    return JsonResponse({'removed': False, 'error': 'Invalid request method'}, status=400)

@login_required
def registered(request):
    registered = Registered.objects.filter(user=request.user)
    reviews = {}
    events = []

    online_ids = []
    offline_ids = []
    attractions_ids = []
    for_visiting_ids = []

    registered_ids = set()
    registered_event_map = {}

    for reg in registered:
        if reg.online:
            events.append(reg.online)
            online_ids.append(reg.online.id)
            registered_ids.add(reg.online.id)
            registered_event_map[reg.online.id] = reg.id
        elif reg.offline:
            events.append(reg.offline)
            offline_ids.append(reg.offline.id)
            registered_ids.add(reg.offline.id)
            registered_event_map[reg.offline.id] = reg.id
        elif reg.attractions:
            events.append(reg.attractions)
            attractions_ids.append(reg.attractions.id)
            registered_ids.add(reg.attractions.id)
            registered_event_map[reg.attractions.id] = reg.id
        elif reg.for_visiting:
            events.append(reg.for_visiting)
            for_visiting_ids.append(reg.for_visiting.id)
            registered_ids.add(reg.for_visiting.id)
            registered_event_map[reg.for_visiting.id] = reg.id

    for event in events:
        content_type = ContentType.objects.get_for_model(event)
        reviews[event.unique_id] = Review.objects.filter(content_type=content_type, object_id=event.id)

    favorites_online_id = Favorite.objects.filter(user=request.user, online_id__in=online_ids).values_list('online_id', 'id')
    favorites_offline_id = Favorite.objects.filter(user=request.user, offline_id__in=offline_ids).values_list('offline_id', 'id')
    favorites_attractions_id = Favorite.objects.filter(user=request.user, attractions_id__in=attractions_ids).values_list('attractions_id', 'id')
    favorites_for_visiting_id = Favorite.objects.filter(user=request.user, for_visiting_id__in=for_visiting_ids).values_list('for_visiting_id', 'id')

    favorites_dict = {
        'online': {item[0]: item[1] for item in favorites_online_id},
        'offline': {item[0]: item[1] for item in favorites_offline_id},
        'attractions': {item[0]: item[1] for item in favorites_attractions_id},
        'for_visiting': {item[0]: item[1] for item in favorites_for_visiting_id},
    }

    favorites_online = Favorite.objects.filter(user=request.user, online_id__in=online_ids).select_related('online')
    favorites_offline = Favorite.objects.filter(user=request.user, offline_id__in=offline_ids).select_related('offline')
    favorites_attractions = Favorite.objects.filter(user=request.user, attractions_id__in=attractions_ids).select_related('attractions')
    favorites_for_visiting = Favorite.objects.filter(user=request.user, for_visiting_id__in=for_visiting_ids).select_related('for_visiting')

    liked_slugs = [
    favorite.online.slug for favorite in favorites_online if favorite.online
] + [
    favorite.offline.slug for favorite in favorites_offline if favorite.offline
] + [
    favorite.attractions.slug for favorite in favorites_attractions if favorite.attractions
] + [
    favorite.for_visiting.slug for favorite in favorites_for_visiting if favorite.for_visiting
]
    

    reviews_avg = {}
    for event in events:
        content_type = ContentType.objects.get_for_model(event)
        avg_rating = Review.objects.filter(
            content_type=content_type,
            object_id=event.id,
            rating__isnull=False
        ).aggregate(Avg('rating'))['rating__avg']
        reviews_avg[event.id] = round(avg_rating, 1) if avg_rating else 0

    print(f'Reviews: {reviews_avg}')


    context = {
        'registered': registered,
        'registered_ids': registered_ids, #для кнопки регистрация/отмена
        'registered_map': registered_event_map, #для data-event-id в регистрации
        'reviews': reviews,
        'name_page': 'Зарегистрированные',
        'liked': liked_slugs,
        'favorites': favorites_dict,
        'reviews_avg': reviews_avg,
    }
    return render(request, 'bookmarks/registered.html', context)


@staff_member_required
def get_event_choices(request):
    event_type = request.GET.get('event_type')
    events = []

    if event_type == 'online':
        events = Events_online.objects.filter(events_admin=request.user)
    elif event_type == 'offline':
        events = Events_offline.objects.filter(events_admin=request.user)
    elif event_type == 'attractions':
        events = Attractions.objects.filter(events_admin=request.user)
    elif event_type == 'for_visiting':
        events = Events_for_visiting.objects.filter(events_admin=request.user)

    event_data = [{'id': event.id, 'name': event.name} for event in events]
    return JsonResponse(event_data, safe=False)

@staff_member_required
def reports(request):
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, "У вас нет прав для доступа к отчетам.")
        return redirect('home')

    # Обработка POST запросов для форм
    if request.method == 'POST':
        form_type = request.POST.get('form_type')
        
        if form_type == 'export_participants':
            event_type = request.POST.get('event_type')
            form = ExportParticipantsForm(request.POST, event_type=event_type, user=request.user)
            if form.is_valid():
                event = form.cleaned_data['event']
                export_all = form.cleaned_data['export_all']
                selected_users = list(form.cleaned_data.get('selected_users', []))

                if not request.user.is_superuser and request.user not in event.events_admin.all():
                    messages.error(request, "У вас нет прав для выгрузки участников этого мероприятия.")
                    return redirect('bookmarks:reports')

                filter_kwargs = {f"{form.cleaned_data['event_type']}": event}
                registrations = Registered.objects.filter(**filter_kwargs).select_related('user')

                if not registrations.exists():
                    messages.error(request, "Нет зарегистрированных участников для выбранного мероприятия.")
                    return redirect('bookmarks:reports')

                def user_full_name(u):
                    parts = [u.last_name or '', u.first_name or '', u.middle_name or '']
                    return ' '.join([p for p in parts if p]).strip()

                # Преобразуем QuerySet в список для фильтрации и сортировки
                registrations = list(registrations)

                if not export_all and selected_users:
                    user_ids = set(u.id for u in selected_users)
                    registrations = [r for r in registrations if r.user_id in user_ids]

                # Сортируем по фамилии, имени, отчеству в алфавитном порядке
                registrations = sorted(registrations, key=lambda r: (
                    (r.user.last_name or '').lower(),
                    (r.user.first_name or '').lower(), 
                    (r.user.middle_name or '').lower()
                ))

                participants = [
                    {"idx": i + 1, "full_name": user_full_name(r.user)}
                    for i, r in enumerate(registrations)
                ]

                if DocxTemplate is None:
                    messages.error(request, "Сервер не поддерживает генерацию DOCX (docxtpl не установлен).")
                    return redirect('bookmarks:reports')

                template_path = os.path.join(settings.BASE_DIR, 'users', 'docx_forms', 'Заявка на проход на мероприятия.docx')
                if not os.path.exists(template_path):
                    messages.error(request, f"Шаблон не найден: {template_path}")
                    return redirect('bookmarks:reports')

                try:
                    doc = DocxTemplate(template_path)
                    context = {
                        'event_name': getattr(event, 'name', ''),
                        'event_category': getattr(event, 'category', ''),
                        'event_date': getattr(event, 'formatted_date_range', lambda: '')() if hasattr(event, 'formatted_date_range') else '',
                        'participants': participants,
                    }
                    doc.render(context)

                    output = BytesIO()
                    doc.save(output)
                    output.seek(0)

                    safe_event = slugify(getattr(event, 'name', 'event'))
                    filename = f"spisok-uchastnikov-{safe_event}.docx"

                    from django.http import HttpResponse
                    response = HttpResponse(
                        output.read(),
                        content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
                    )
                    response['Content-Disposition'] = f'attachment; filename="{filename}"'
                    return response
                except Exception as e:
                    messages.error(request, f"Ошибка при обработке шаблона: {str(e)}. Проверьте правильность тегов в DOCX файле.")
                    return redirect('bookmarks:reports')
            else:
                messages.error(request, "Ошибка в форме экспорта участников")
        
        elif form_type == 'export_logistics':
            form = ExportLogisticsForm(request.POST, user=request.user)
            if form.is_valid():
                event = form.cleaned_data['event']
                if not request.user.is_superuser and request.user not in event.events_admin.all():
                    messages.error(request, "У вас нет прав для выгрузки логистики этого мероприятия.")
                    return redirect('bookmarks:reports')
                
                # Логика экспорта логистики (копируем из export_logistics)
                from events_available.models import EventLogistics
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                from openpyxl.utils import get_column_letter
                from django.utils import timezone
                from django.http import HttpResponse
                from django.utils.text import slugify as dj_slugify

                def _build_xlsx_for_event(event):
                    qs = EventLogistics.objects.filter(event=event).select_related('user').order_by('user__last_name', 'user__first_name')
                    if not qs.exists():
                        return None
                    wb = Workbook(); ws = wb.active; ws.title = "Логистика"
                    title = f"Логистика по мероприятию \"{event.name}\""
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
                    cell = ws.cell(row=1, column=1, value=title); cell.font = Font(size=14, bold=True); cell.alignment = Alignment(horizontal='center')
                    headers = ["ФИО","Дата/время прилёта","Рейс прилёта","Аэропорт прилёта","Дата/время отлёта","Рейс отлёта","Аэропорт отлёта","Нужен трансфер","Гостиница","Встречающий"]
                    ws.append(headers)
                    head_fill = PatternFill(start_color="FFEEF5FF", end_color="FFEEF5FF", fill_type="solid")
                    thin = Side(style='thin', color='FFBBBBBB'); border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    for col_idx in range(1, len(headers)+1):
                        c = ws.cell(row=2, column=col_idx); c.font = Font(bold=True); c.fill = head_fill; c.border = border; c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    row = 3
                    for rec in qs:
                        fio = ' '.join(filter(None, [rec.user.last_name, rec.user.first_name, getattr(rec.user, 'middle_name', '') or None]))
                        arr_dt = rec.arrival_datetime.astimezone(timezone.get_current_timezone()).strftime('%d.%m.%Y %H:%M') if rec.arrival_datetime else ''
                        dep_dt = rec.departure_datetime.astimezone(timezone.get_current_timezone()).strftime('%d.%m.%Y %H:%M') if rec.departure_datetime else ''
                        row_values = [fio, arr_dt, rec.arrival_flight_number or '', rec.arrival_airport or '', dep_dt, rec.departure_flight_number or '', rec.departure_airport or '', "Да" if rec.transfer_needed else "Нет", rec.hotel_details or '', rec.meeting_person or '']
                        ws.append(row_values)
                        for col_idx in range(1, len(headers)+1):
                            c = ws.cell(row=row, column=col_idx); c.border = border; c.alignment = Alignment(vertical='top', wrap_text=True)
                        row += 1
                    # Автофильтр по шапке
                    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{ws.max_row if ws.max_row>2 else 2}"
                    for col_idx in range(1, len(headers)+1):
                        max_len = 0; col_letter = get_column_letter(col_idx)
                        for r in range(2, ws.max_row+1):
                            val = ws.cell(row=r, column=col_idx).value; max_len = max(max_len, len(str(val)) if val else 0)
                        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
                    filename = f"logistika-{dj_slugify(event.name)}.xlsx"
                    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = f'attachment; filename="{filename}"'
                    wb.save(response)
                    return response

                resp = _build_xlsx_for_event(event)
                if resp is None:
                    messages.warning(request, "Логистика отсутствует для выбранного мероприятия.")
                    return redirect('bookmarks:reports')
                return resp
            else:
                messages.error(request, "Ошибка в форме экспорта логистики")

    # Инициализация форм для отображения
    export_participants_form = ExportParticipantsForm(user=request.user)
    export_logistics_form = ExportLogisticsForm(user=request.user)
    
    # Подсчет участников для логистики
    selected_event = export_logistics_form['event'].value()
    participants_count = 0
    if selected_event:
        try:
            participants_count = Registered.objects.filter(offline_id=selected_event).count()
        except Exception:
            participants_count = 0

    context = {
        'export_participants_form': export_participants_form,
        'export_logistics_form': export_logistics_form,
        'participants_count': participants_count,
    }
    
    return render(request, 'bookmarks/reports_page.html', context)


@staff_member_required
def send_message_to_participants(request):
    if not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, "У вас нет прав для отправки сообщений.")
        return redirect('home')

    if request.method == 'POST':
        event_type = request.POST.get('event_type')
        logger.info(f"Получены POST-данные: {request.POST}")
        logger.info(f"Выбранный тип мероприятия: {event_type}")
        
        form = SendMessageForm(request.POST, event_type=event_type, user=request.user)
        if form.is_valid():
            logger.info("Форма валидна")
            event = form.cleaned_data['event']
            message = form.cleaned_data['message']
            send_to_all = form.cleaned_data['send_to_all']
            selected_users = form.cleaned_data.get('selected_users', [])

            logger.info(f"Очищенные данные формы: event={event}, send_to_all={send_to_all}, selected_users={[user.username for user in selected_users]}")
            
            if not event:
                messages.error(request, "Пожалуйста, выберите мероприятие.")
                return redirect('bookmarks:send_message_to_participants')

            # Проверяем права доступа к мероприятию
            if not request.user.is_superuser and request.user not in event.events_admin.all():
                messages.error(request, "У вас нет прав для отправки сообщений участникам этого мероприятия.")
                return redirect('bookmarks:send_message_to_participants')

            # Получаем всех зарегистрированных участников для данного мероприятия
            filter_kwargs = {f"{event_type}": event}
            registered_users = Registered.objects.filter(**filter_kwargs)
            
            logger.info(f"Найдено {registered_users.count()} участников для мероприятия {event.name}")

            if not registered_users.exists():
                messages.error(request, "Нет зарегистрированных участников для выбранного мероприятия.")
                return redirect('bookmarks:send_message_to_participants')

            # Отправляем сообщения
            sent_count = 0
            for registration in registered_users:
                should_send = send_to_all or (selected_users and registration.user in selected_users)
                logger.info(f"Проверка отправки для {registration.user.username}: send_to_all={send_to_all}, selected_users_empty={not selected_users}, в selected_users={registration.user in selected_users}, итоговое решение={should_send}")
                
                if should_send:
                    if registration.user.telegram_id:
                        try:
                            logger.info(f"Отправка сообщения пользователю {registration.user.username} (Telegram ID: {registration.user.telegram_id})")
                            response = send_notification_with_toggle(
                                telegram_id=registration.user.telegram_id,
                                message=message,
                                event_id=event.unique_id,
                                notifications_enabled=True
                            )
                            logger.info(f"Ответ от Telegram API: {response}")
                            if response and response.get('ok'):
                                sent_count += 1
                                logger.info(f"Сообщение успешно отправлено пользователю {registration.user.username}")
                            else:
                                logger.error(f"Ошибка отправки сообщения пользователю {registration.user.username}: {response.get('description', 'Неизвестная ошибка')}")
                        except Exception as e:
                            logger.error(f"Ошибка отправки сообщения пользователю {registration.user.username}: {e}")
                    else:
                        logger.warning(f"У пользователя {registration.user.username} нет Telegram ID")

            if sent_count > 0:
                messages.success(request, f"Сообщения успешно отправлены {sent_count} участникам.")
            else:
                messages.warning(request, "Не удалось отправить ни одного сообщения.")
            return redirect('bookmarks:send_message_to_participants')
        else:
            logger.error(f"Ошибки формы: {form.errors}")
            #messages.error(request, f"Ошибка в форме: {form.errors}")
            messages.error(request, f"Ошибка при отправке")

    else:
        form = SendMessageForm(user=request.user)

    return render(request, 'bookmarks/send_message.html', {'form': form})

@staff_member_required
def get_event_participants(request):
    event_id = request.GET.get('event_id')
    event_type = request.GET.get('event_type')
    
    if not event_id or not event_type:
        return JsonResponse({'error': 'Missing parameters'}, status=400)
        
    model_map = {
        'online': Events_online,
        'offline': Events_offline,
        'attractions': Attractions,
        'for_visiting': Events_for_visiting
    }
    
    model = model_map.get(event_type)
    if not model:
        return JsonResponse({'error': 'Invalid event type'}, status=400)
        
    try:
        event = model.objects.get(id=event_id)
        
        # Проверяем права доступа
        if not request.user.is_superuser and request.user not in event.events_admin.all():
            return JsonResponse({'error': 'Access denied'}, status=403)
            
        registered_users = Registered.objects.filter(**{event_type: event}).select_related('user')
        participants = [
            {
                'id': reg.user.id,
                'name': f"{reg.user.last_name} {reg.user.first_name} ({reg.user.username})"
            }
            for reg in registered_users
        ]
        return JsonResponse(participants, safe=False)
    except model.DoesNotExist:
        return JsonResponse({'error': 'Event not found'}, status=404)

